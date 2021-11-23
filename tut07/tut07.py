import pandas as pd
import csv
import os
from openpyxl import Workbook
from openpyxl import load_workbook
import shutil
def feedback_not_submitted():
	course_dict = {}
	feedback_dict = {}
	remaining_dict = {}
	student_info = {}
	remaining_file = os.path.join('course_feedback_remaining.xlsx')
	if os.path.isfile(remaining_file):
		os.remove(remaining_file)
	with open('course_master_dont_open_in_excel.csv','r') as file:
		course_master = csv.DictReader(file)
		for line in course_master:
			ltp_data = line['ltp'].strip().split('-')
			x = 0
			for i in ltp_data:
				if i != '0':
					x+=1
			course_dict[line['subno'].strip()] = {'non_zero':x,'l':ltp_data[0],'t':ltp_data[1],'p':ltp_data[2]}
		# print(course_dict['NSS'])

	with open('course_feedback_submitted_by_students.csv','r') as file:
		feedback = csv.DictReader(file)
		for line in feedback:
			roll_no = line['stud_roll'].strip()
			subno = line['course_code'].strip()
			if roll_no+subno not in feedback_dict.keys():
				feedback_dict[roll_no+subno] = {'l':0,'t':0,'p':0}
			if line['feedback_type'] == '1':
				feedback_dict[roll_no+subno]['l'] = 1
			if line['feedback_type'] == '2':
				feedback_dict[roll_no+subno]['t'] = 1
			if line['feedback_type'] == '3':
				feedback_dict[roll_no+subno]['p'] = 1
		
		# print(feedback_dict['2001CE54'+'NSS'])
	with open('studentinfo.csv','r') as file:
		info = csv.DictReader(file)
		for line in info:
			student_info[line['Roll No']] = {'Name':line['Name'],'email':line['email'],'aemail':line['aemail'],'contact':line['contact']}

	with open('course_registered_by_all_students.csv','r') as file:
		course_register = csv.DictReader(file)
		ans = 0
		for line in course_register:
			roll_no = line['rollno'].strip()
			subno = line['subno'].strip()
			Na = "NA_IN_STUDENTINFO"
			remaining_dict[roll_no+subno] = 1
			l_real = course_dict[line['subno']]['l']
			t_real = course_dict[line['subno']]['t']
			p_real = course_dict[line['subno']]['p']
			if l_real != '0' or t_real != '0' or p_real!='0':
				if roll_no+subno in feedback_dict.keys():
					l_sub = feedback_dict[roll_no+subno]['l']
					t_sub = feedback_dict[roll_no+subno]['t']
					p_sub = feedback_dict[roll_no+subno]['p']
					if float(l_real)>0 and l_sub == 0:
						remaining_dict[roll_no+subno] = 0
					if float(t_real)>0 and t_sub == 0:
						remaining_dict[roll_no+subno] = 0
					if float(p_real)>0 and p_sub == 0:
						remaining_dict[roll_no+subno] = 0
				else:
					remaining_dict[roll_no+subno] = 0

			if remaining_dict[roll_no+subno] == 0:
				if os.path.isfile(remaining_file) is True:
					#If the file is present then loading it and appending data in it.
					wb = load_workbook(remaining_file)
					sheet = wb.active
					if roll_no in student_info.keys():
						sheet.append([roll_no,line['register_sem'],line['schedule_sem'],subno,student_info[roll_no]['Name'],student_info[roll_no]['email'],student_info[roll_no]['aemail'],student_info[roll_no]['contact']]) 
					else:
						sheet.append([roll_no,line['register_sem'],line['schedule_sem'],subno,Na,Na,Na,Na])
					wb.save(remaining_file)
				else:
					#if the file is not present then creating the file and adding the header line.
					wb = Workbook()
					sheet = wb.active
					sheet.append(['rollno','register_sem','schedule_sem','subno','Name','email','aemail','contact'])
					if roll_no in student_info.keys():
						sheet.append([roll_no,line['register_sem'],line['schedule_sem'],subno,student_info[roll_no]['Name'],student_info[roll_no]['email'],student_info[roll_no]['aemail'],student_info[roll_no]['contact']])     
					else:
						sheet.append([roll_no,line['register_sem'],line['schedule_sem'],subno,Na,Na,Na,Na])
					wb.save(remaining_file)
		# for i in remaining_dict:
			# print(i,'->',remaining_dict[i])
		#     if remaining_dict[roll_no+subno]==0:
		#         ans+=1
		# print(ans)

feedback_not_submitted()
