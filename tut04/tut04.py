from typing import ClassVar
import csv
import os
from openpyxl import Workbook
from openpyxl import load_workbook
#All the files have been generated already in the folders, delete the 'output_by_subject'
# and output_individual_roll folders for generating the data again by running the code.
def output_by_subject():
    #Checking if the required folder is available or not, If not then creating the folder. 
    parentdir="output_by_subject"
    if os.path.exists(parentdir)==False:
        os.mkdir(parentdir) 
    #Opening the main data file
    with open('regtable_old.csv','r') as file:
        reader = csv.DictReader(file)
        for x in reader:
            #Checking if the required subject file is present or not
            targetfile=os.path.join(parentdir,x['subno']+".xlsx")
            if os.path.isfile(targetfile) is True:
                #If the file is present then loading it and appending data in it.
                wb = load_workbook(targetfile)
                sheet = wb.active
                sheet.append([x['rollno'],x['register_sem'],x['subno'],x['sub_type']])
                wb.save(targetfile)
            else:
                #if the file is not present then creating the file and adding the header line.
                wb = Workbook()
                sheet = wb.active
                sheet.append(['rollno','register_sem','subno','sub_type'])
                sheet.append([x['rollno'],x['register_sem'],x['subno'],x['sub_type']])     
                wb.save(targetfile)
    return

def output_individual_roll():
    #Checking if the required folder is available or not, If not then creating the folder.
    parentdir="output_individual_roll"
    if os.path.exists(parentdir)==False:
        os.mkdir(parentdir) 
    #Opening the main data file
    with open('regtable_old.csv','r') as file:
        reader = csv.DictReader(file)
        for x in reader:
            #Checking if the required subject file is present or not
            targetfile=os.path.join(parentdir,x['rollno']+".xlsx")
            if os.path.isfile(targetfile) is True:
                #If the file is present then loading it and appending data in it.
                wb = load_workbook(targetfile)
                sheet = wb.active
                sheet.append([x['rollno'],x['register_sem'],x['subno'],x['sub_type']])
                wb.save(targetfile)
            else:
                #if the file is not present then creating the file and adding the header line.
                wb = Workbook()
                sheet = wb.active
                sheet.append(['rollno','register_sem','subno','sub_type'])
                sheet.append([x['rollno'],x['register_sem'],x['subno'],x['sub_type']])     
                wb.save(targetfile)
    return

output_individual_roll()
output_by_subject()
