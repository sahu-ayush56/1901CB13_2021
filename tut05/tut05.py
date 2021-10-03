from os import name
import os.path
import shutil
import csv
from typing import Dict
from openpyxl import Workbook
from openpyxl import load_workbook

def generate_marksheet():
    direc = "output"
    dirpath = os.path.join(direc)
    #If output folder is present then removing it and then making it again for avoiding readding of data
    if os.path.isdir(dirpath) == True:
        shutil. rmtree(dirpath)
    os.mkdir(direc)
    #Mde four dictionaries for storing the data from names-roll.csv and subjects_master.csv
    namedict = {}
    subnames = {}
    subltp = {}
    Discipline = {}
    #storing data from names-roll.csv
    with open('names-roll.csv','r') as file:
        namesroll = csv.DictReader(file)
        for i in namesroll:
            namedict[i['Roll'].strip()] = i['Name']
            D = ""
            for char in i['Roll']:
                D += char if ((ord(char)>=65 and ord(char)<=90) or (ord(char) >= 97 and ord(char) <= 122)) else ""
            Discipline[i['Roll'].strip()] = D;
    #storing data from subjects_master.csv          
    with open('subjects_master.csv','r') as file:
        subjects = csv.DictReader(file)
        for i in subjects:
            subnames[i['subno']] = i['subname']
            subltp[i['subno']] = i['ltp']
    #Now iterating on grades.csv
    with open('grades.csv','r') as file:
        grades = csv.DictReader(file)
        for i in grades:
            #Checking if the roll_no.xlsx file is present in output folder or not
            roll = i['Roll'].strip()
            targetfile = os.path.join(direc,roll+".xlsx")

            if os.path.isfile(targetfile) == True:
                #As the file is present, so laoding the workbook
                wb = load_workbook(targetfile)
                sheet = wb.active
                Sem = i['Sem']
                semsheet = "Sem"+Sem
                #Checking if a sheet with Sem# is present or not.
                if semsheet in wb.sheetnames:
                    #As is it is present, so just appending the data from grade.csv in the sheet.
                    si = wb[semsheet].max_row
                    wb[semsheet].append([si,i['SubCode'].strip(),subnames[i['SubCode'].strip()],subltp[i['SubCode'].strip()],i['Credit'],i['Sub_Type'].strip(),i['Grade'].strip()])
                    wb.save(targetfile)
                else:
                    #As it is not present, so creating a new Sheet in the required roll_no.xlsx file and adding the initial contents
                    wb.create_sheet(index = int(Sem), title = semsheet)
                    wb[semsheet].append(['SI No.','Subject No.','Subject Name','L-T-P','Credit','Subject Type','Grade'])
                    si = 1
                    wb[semsheet].append([si,i['SubCode'].strip(),subnames[i['SubCode'].strip()],subltp[i['SubCode'].strip()],i['Credit'],i['Sub_Type'].strip(),i['Grade'].strip()])
                    wb.save(targetfile)
            else:
                #As the file is not present, so making a new Workbook
                wb = Workbook()
                sheet = wb.active
                sheet.title = "Overall"
                Sem = i['Sem']
                semsheet = "Sem"+Sem
                wb.create_sheet(index = int(Sem), title = semsheet)
                wb[semsheet].append(['SI No.','Subject No.','Subject Name','L-T-P','Credit','Subject Type','Grade'])
                si = 1
                wb[semsheet].append([si,i['SubCode'].strip(),subnames[i['SubCode'].strip()],subltp[i['SubCode'].strip()],i['Credit'],i['Sub_Type'].strip(),i['Grade'].strip()])
                wb.save(targetfile)
    #Now in every roll_no file present, we have made every sem sheet and added its content, but we have not added content in overall sheet, so iterating over names-roll.csv for going roll no wise to add overall sheet contents
    with open('names-roll.csv','r') as file:
        #A dictionary for taking care of grades 
        gradedict = {'AA':10,'AB':9,'BB':8,'BC':7,'CC':6,'CD':5,'DD':4,'F':0,'I':0,'DD*':4,'F*':0}

        finalprf = csv.DictReader(file)
        for f in finalprf:
            flag = 0
            #storing file path of the roll_no.xlsx in target variable
            target = os.path.join(direc,f['Roll'].strip()+".xlsx")
            if os.path.isfile(target) == True:
                wb = load_workbook(target)
                sheet = wb.active
                #Firsty adding simple contents like Roll no, name of student and Discipline
                wb["Overall"].append(['Roll No.',f['Roll'].strip()])
                wb["Overall"].append(['Name of Student',namedict[f['Roll'].strip()]])
                wb["Overall"].append(['Discipline',Discipline[f['Roll'].strip()]])
                #Now making lists of other data to be appended after calculation.
                semlist = ['Semester No.']
                semcredit = ['Semester Wise Credit Taken']
                totalcredits = ['Total Credits Taken']
                spi = ['SPI']
                cpi = ['CPI']
                creditssum = 0
                #Now we will be iterating every sem sheet of the roll_no we are currently on.
                for n in wb.sheetnames:
                    #Now this condition is for avoiding interence with Overall sheet, we will be adding data in Overall sheet, but firstly we need to collect the data.
                    if flag==1:
                        #num is making extracting sem no.
                        num = n[3:]
                        semlist.append(num)
                        sheet = wb[n]
                        creditpersem = 0
                        credit = []
                        gradepersubject = []
                        #Now iterating in the sheet we are currently on.
                        for i in range(1,sheet.max_column+1):
                            
                            for j in range(1,sheet.max_row+1):
                                cell_obj = sheet.cell(row=j, column = i)
                                if(i==5 and j > 1):
                                    #Credit column data
                                    creditpersem+=int(cell_obj.value)
                                    credit.append(int(cell_obj.value))
                                if(i == 7 and j > 1):
                                    #Grade column data 
                                    cell_obj = sheet.cell(row=j, column = i)
                                    gradepersubject.append(cell_obj.value)
                        sum = 0           
                        for k in range(0,len(credit)):
                            sum += credit[k]*gradedict[gradepersubject[k]]  
                        #spi calculation
                        spisem = round(sum/creditpersem,2)
                        #adding credits after every sem
                        creditssum += creditpersem
                        #Now appending the data
                        semcredit.append(creditpersem)
                        spi.append(spisem)
                        totalcredits.append(creditssum)
                    flag = 1
                cpi.append(spi[1])
                #cpi calculation
                for i in range(1,len(spi)-1):
                    numerator = cpi[i]*totalcredits[i] + spi[i+1]*semcredit[i+1]
                    denominator = totalcredits[i+1]
                    cpipersem = round(numerator/denominator,2)
                    cpi.append(cpipersem)
                #Now adding the collected data in the Overall sheet
                wb["Overall"].append(semlist)  
                wb["Overall"].append(semcredit)
                wb["Overall"].append(spi)
                wb["Overall"].append(totalcredits)
                wb["Overall"].append(cpi)
                # for n in wb.sheetnames:
                   
                #Saving the roll_no file
                wb.save(target)

    return


generate_marksheet()


